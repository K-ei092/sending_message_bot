import time
from random import randint

from openpyxl import Workbook, load_workbook

from sqlalchemy.ext.asyncio.engine import AsyncEngine
from sqlalchemy import insert, select, CursorResult, delete

from database.tables import users as users_table
from lexicon.lexicon import LEXICON


async def add_new_user(db_engine: AsyncEngine,
                       telegram_id: int,
                       telegram_name: str | None,
                       first_name: str | None,
                       last_name: str | None
                       ) -> None:
    stmt = insert(users_table).values(
        telegram_id=telegram_id,
        telegram_name=telegram_name,
        first_name=first_name,
        last_name=last_name,
        exercises_counter=0
    )
    async with db_engine.connect() as conn:
        await conn.execute(stmt)
        await conn.commit()


def _create_excel_for_admin(result: CursorResult) -> str:
    timestr = time.strftime('%Y.%m.%d_%H-%M')
    name_file = f'Users_bot({timestr}).xlsx'
    count_row = 2
    result_number = 1

    wb = Workbook()
    wb.save(name_file)
    wb = load_workbook(name_file)
    ws = wb.active

    ws.cell(row=1, column=1, value='№')
    ws.cell(row=1, column=2, value='telegram_id')
    ws.cell(row=1, column=3, value='telegram_name')
    ws.cell(row=1, column=4, value='first_name')
    ws.cell(row=1, column=5, value='last_name')
    ws.cell(row=1, column=6, value='добавлен в базу')
    ws.cell(row=1, column=7, value='количество баллов')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20

    wb.save(name_file)

    for res in result:
        ws.cell(row=count_row, column=1, value=f'{result_number}')
        ws.cell(row=count_row, column=2, value=res[0])
        ws.cell(row=count_row, column=3, value=f'@{res[1]}' if res[1] else None)
        ws.cell(row=count_row, column=4, value=res[2])
        ws.cell(row=count_row, column=5, value=res[3])
        ws.cell(row=count_row, column=6, value=res[6])
        ws.cell(row=count_row, column=7, value=res[8])
        count_row += 1
        result_number += 1

    wb.save(name_file)

    return name_file


async def get_file_db(db_engine: AsyncEngine) -> str:
    stmt = select("*").select_from(users_table)
    async with db_engine.connect() as conn:
        result = await conn.execute(stmt)

    name_file = _create_excel_for_admin(result=result)

    return name_file


async def check_user_id(db_engine: AsyncEngine, user_id: int) -> bool:
    stmt = (
        select(users_table.c.telegram_id)
        .where(users_table.c.telegram_id == user_id)
    )
    async with db_engine.connect() as conn:
        results = await conn.execute(stmt)
        if not results.first():
            return False
        else:
            return True


async def reset_user_count(db_engine: AsyncEngine, user_id: int) -> None:
    stmt = (
        users_table.update()
        .where(users_table.c.telegram_id == user_id)
        .values(exercises_counter=0)
    )
    async with db_engine.connect() as conn:
        await conn.execute(stmt)
        await conn.commit()


async def get_all_from_db(db_engine: AsyncEngine) -> CursorResult:
    stmt = select("*").select_from(users_table)
    async with db_engine.connect() as conn:
        users = await conn.execute(stmt)

    return users


async def get_all_user_id(db_engine: AsyncEngine) -> list:
    stmt = select(users_table.c.telegram_id).select_from(users_table)
    async with db_engine.connect() as conn:
        result = await conn.execute(stmt)

    list_id = []

    for tuple_user in result:
        list_id.append(tuple_user[0])

    return list_id


async def delete_user_database(db_engine: AsyncEngine, user_id: int) -> bool:
    stmt = (
        delete(users_table)
        .where(users_table.c.telegram_id == user_id)
    )
    async with db_engine.connect() as conn:
        await conn.execute(stmt)
        await conn.commit()
        return True


async def change_exercise_state(db_engine: AsyncEngine, user_id: int, state: bool) -> None:
    stmt = (
        users_table.update()
        .where(users_table.c.telegram_id == user_id)
        .values(exercise_state=state)
    )
    async with db_engine.connect() as conn:
        await conn.execute(stmt)
        await conn.commit()


async def increase_record(db_engine: AsyncEngine, user_id: int) -> None:
    stmt = (
        users_table.update()
        .where(users_table.c.telegram_id == user_id)
        .values(record_counter=users_table.c.record_counter + 1)
    )
    async with db_engine.connect() as conn:
        await conn.execute(stmt)
        await conn.commit()


async def get_statistic(db_engine: AsyncEngine, user_id: int) -> str:
    stmt = select(users_table.c.first_name, users_table.c.record_counter) \
        .order_by(users_table.c.record_counter.desc())

    async with db_engine.connect() as conn:
        result = await conn.execute(stmt)

    rating_text = f"<u>{LEXICON['rating_handlers'][randint(1, len(LEXICON['rating_handlers']) - 1)]}</u>:\n"
    current_rank = 0
    current_score = None
    current_names = []

    for row in result:
        if row[1] != current_score and current_rank <= 10:
            if current_names:
                rating_text += f"{current_rank} место: <b>{', '.join(current_names)}</b> с рейтингом - {current_score}\n"
                current_names = []
            current_score = row[1]
            current_rank += 1
        current_names.append(row[0])

    stmt = select(users_table.c.record_counter).where(users_table.c.telegram_id == user_id)

    async with db_engine.connect() as conn:
        results = await conn.execute(stmt)
        results = results.fetchall()

    if results:
        user_rating = results[0][0]
        rating_text += f'\n<b>Ваш рейтинг - {user_rating}</b>'
    else:
        rating_text += f'\n<b>У вас нет рейтинга. Перезапустите бот</b>'

    return rating_text


async def get_false_state_exercise(db_engine: AsyncEngine) -> CursorResult:
    stmt = (
        select(users_table.c.telegram_id)
        .where(users_table.c.exercise_state == False)
    )
    async with db_engine.connect() as conn:
        results = await conn.execute(stmt)

    return results


async def get_state_exercise(db_engine: AsyncEngine, user_id: int) -> CursorResult:
    stmt = (
        select(users_table.c.exercise_state)
        .where(users_table.c.telegram_id == user_id)
    )
    async with db_engine.connect() as conn:
        results = await conn.execute(stmt)

    return results
