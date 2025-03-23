import logging
import time
from random import randint
from typing import Any

from openpyxl import Workbook, load_workbook

from sqlalchemy.ext.asyncio.engine import AsyncEngine
from sqlalchemy import insert, select, CursorResult, delete, and_, Row
from sqlalchemy.sql import Insert, Delete, Update

from database.tables import users as users_table
from lexicon.lexicon import LEXICON


logger = logging.getLogger(__name__)


class WorkerDB:
    def __init__(self, db_engine: AsyncEngine):
        self.db_engine = db_engine

    async def _execute(self, stmt) -> CursorResult[Any] | None:
        async with self.db_engine.connect() as conn:
            coroutine = await conn.execute(stmt)
            logger.info('----------------------------------------')
            logger.info(type(stmt))
            logger.info('----------------------------------------')
            if isinstance(stmt, (Insert, Delete, Update)):
                await conn.commit()
                logger.info('commit database')
                return None
            logger.info('return coroutine')
            return coroutine

    async def add_user(self, telegram_id: int, telegram_name: str | None, first_name: str | None, last_name: str | None) -> None:
        stmt = insert(users_table).values(telegram_id=telegram_id, telegram_name=telegram_name, first_name=first_name, last_name=last_name)
        await self._execute(stmt)

    async def change_after_submitting_exercise(self, user_id: int, exercises_counter: int, completed_exercises: str,
                                               exercise_id: int, exercise_submission_status: bool = True) -> None:
        stmt = (users_table.update().where(users_table.c.telegram_id == user_id).
                        values(exercises_counter=exercises_counter + 1,                                                  # увеличиваем счетчик "30 заданий" на 1
                               completed_exercises=completed_exercises + ' ' + str(exercise_id),                         # добавляем к строке отправленных заданий новое
                               exercise_submission_status=exercise_submission_status))                                   # меняем состояние "задание отправлено" на - "да"
        await self._execute(stmt)

    async def change_exercise_state(self, user_id: int, state: bool) -> None:
        stmt = users_table.update().where(users_table.c.telegram_id == user_id).values(exercise_state=state)             # True - задание выполнено лайк / False - ожидание лайка
        await self._execute(stmt)

    async def change_exercise_submission_status(self, user_id: int, state: bool) -> None:
        stmt = users_table.update().where(users_table.c.telegram_id == user_id).values(exercise_submission_status=state) # задание не отправлено / отправлено
        await self._execute(stmt)

    async def check_before_send_exercise(self, user_id: int) -> Row | None:
        stmt = (select("*").where(and_(users_table.c.telegram_id == user_id,
                                        users_table.c.exercises_counter < 30,                                           # если счетчик "30 отправленных" меньше (30 а не 31, потому что после 29 еще + 1 после отправки)
                                        users_table.c.exercise_state == True,                                           # если пользователь поставил лайк
                                        users_table.c.exercise_submission_status == False)))                            # если новое задание ранее не отправлено
        users = await self._execute(stmt)
        return users.first()

    async def check_before_send_gratitude(self, user_id: int) -> Row | None:
        stmt = (select("*").where(and_(users_table.c.telegram_id == user_id,
                                        users_table.c.exercises_counter < 31,                                           # если счетчик "30 отправленных" меньше
                                        users_table.c.exercise_state == True,                                           # если пользователь поставил лайк
                                        users_table.c.exercise_submission_status == True)))                             # если задание было отправлено
        users = await self._execute(stmt)
        return users.first()

    async def check_user_id(self, user_id: int) -> bool:
        stmt = select(users_table.c.telegram_id).where(users_table.c.telegram_id == user_id)
        result = await self._execute(stmt)
        if not result.first():
            return False
        return True

    async def delete_user(self, user_id: int) -> None:
        stmt = delete(users_table).where(users_table.c.telegram_id == user_id)
        await self._execute(stmt)

    async def get_all_from_db(self) -> CursorResult:
        stmt = select("*").select_from(users_table)
        users = await self._execute(stmt)
        return users

    async def get_all_tg_id(self) -> list:
        stmt = select(users_table.c.telegram_id).select_from(users_table)
        coroutine = await self._execute(stmt)
        ids = self._make_list(coroutine)
        return ids

    async def get_db(self) -> str:
        stmt = select("*").select_from(users_table)
        cursor = await self._execute(stmt)
        name_file = self._create_excel(cursor=cursor)
        return name_file

    async def get_state_exercise(self, user_id: int) -> CursorResult:
        stmt = select(users_table.c.exercise_state).where(users_table.c.telegram_id == user_id)
        results = await self._execute(stmt)
        return results

    async def get_statistic(self, user_id: int) -> str:
        stmt = select(users_table.c.first_name, users_table.c.record_counter).order_by(users_table.c.record_counter.desc())
        coroutine = await self._execute(stmt)
        rating_text = await self._make_rating_text(coroutine, user_id)
        return rating_text

    async def get_unfulfilled_state_exercise(self) -> CursorResult:
        stmt = select(users_table.c.telegram_id).where(users_table.c.exercise_state == False)
        results = await self._execute(stmt)
        return results

    async def increase_record(self, user_id: int) -> None:
        stmt = users_table.update().where(users_table.c.telegram_id == user_id).values(record_counter=users_table.c.record_counter + 1)     # увеличивает счетчик "всего выполнено заданий" + 1
        await self._execute(stmt)

    async def inflate_exercises_counter(self, user_id: int) -> None:
        stmt = users_table.update().where(users_table.c.telegram_id == user_id).values(exercises_counter=32)             # завышение счетчика "30 заданий" - уход от фильтра
        await self._execute(stmt)

    async def reset_exercises_count(self, user_id: int) -> None:
        stmt = users_table.update().where(users_table.c.telegram_id == user_id).values(exercises_counter=0)              # обнулить счетчик 30 отправленных заданий
        await self._execute(stmt)

    async def _make_rating_text(self, coroutine, user_id) -> str:
        rating_text = f"<u>{LEXICON['rating_handlers'][randint(1, len(LEXICON['rating_handlers']) - 1)]}</u>:\n"
        current_rank = 0
        current_score = None
        current_names = []
        for row in coroutine:
            if row[1] != current_score and current_rank <= 10:
                if current_names:
                    rating_text += f"{current_rank} место: <b>{', '.join(current_names)}</b> с рейтингом - {current_score}\n"
                    current_names = []
                current_score = row[1]
                current_rank += 1
            current_names.append(row[0])
        # if current_names:
        #     rating_text += f"{current_rank} место: <b>{', '.join(current_names)}</b> с рейтингом - {current_score}\n"
        stmt = select(users_table.c.record_counter).where(users_table.c.telegram_id == user_id)
        coroutine = await self._execute(stmt)
        results = coroutine.fetchall()
        if results:
            user_rating = results[0][0]
            rating_text += f'\n<b>Ваш рейтинг - {user_rating}</b>'
        else:
            rating_text += f'\n<b>У вас нет рейтинга. Перезапустите бот</b>'
        return rating_text

    @staticmethod
    def _make_list(coroutine):
        list_id = []
        for tuple_user in coroutine:
            list_id.append(tuple_user[0])
        return list_id

    @staticmethod
    def _create_excel(cursor: CursorResult) -> str:
        create_time = time.strftime('%Y.%m.%d_%H-%M')
        name_file = f'Users_bot({create_time}).xlsx'
        count_row = 2
        result_number = 1

        wb = Workbook()
        wb.save(name_file)
        wb = load_workbook(name_file)
        ws = wb.active

        ws.cell(row=1, column=1, value='№')
        ws.cell(row=1, column=2, value='telegram_id')
        ws.cell(row=1, column=3, value='telegram_name')
        ws.cell(row=1, column=4, value='first_name')
        ws.cell(row=1, column=5, value='last_name')
        ws.cell(row=1, column=6, value='добавлен в базу')
        ws.cell(row=1, column=7, value='балл')

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15

        wb.save(name_file)

        for res in cursor:
            ws.cell(row=count_row, column=1, value=f'{result_number}')
            ws.cell(row=count_row, column=2, value=res[0])
            ws.cell(row=count_row, column=3, value=f'@{res[1]}' if res[1] else None)
            ws.cell(row=count_row, column=4, value=res[2])
            ws.cell(row=count_row, column=5, value=res[3])
            ws.cell(row=count_row, column=6, value=res[6])
            ws.cell(row=count_row, column=7, value=res[8])
            count_row += 1
            result_number += 1

        wb.save(name_file)
        return name_file


